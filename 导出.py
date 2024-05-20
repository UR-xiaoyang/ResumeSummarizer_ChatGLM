# 导出为Excel表
# 表头为：姓名，年龄，学位，毕业学校，工龄，项目经历，总体分析，AI建议

from openpyxl import Workbook
def 输出Excel(工程文件夹,data):
    # 检查
    longest_key = max(data, key=lambda x: len(data[x]) if isinstance(data[x], list) else 1)
    # 补空位
    for i in data:
        if len(data[i]) < len(data[longest_key]):
            for j in range(len(data[longest_key]) - len(data[i])):
                data[i].append("Null")
    # 将字典导出为Excel表
    # 创建一个新的工作表
    工作表 = Workbook()
    # 设为活动
    活动的工作表 = 工作表.active
    # 写入表头
    for 列号, 列名 in enumerate(data.keys(), start=1):
        活动的工作表.cell(row=1, column=列号, value=列名)
    # 写入数据
    for 行号 in range(2, len(data[longest_key]) + 2):
        for 索引,键名 in enumerate(data.keys()):
            活动的工作表.cell(row=行号, column=索引 + 1, value=data[键名][行号 - 2])



    filename = f"{工程文件夹}/简历分析.xlsx"
    工作表.save(filename)
    print(f"数据已导出到Excel文件：{filename}")
